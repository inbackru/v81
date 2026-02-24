"""
Comprehensive import script for all Excel data files
Imports all Excel files uploaded by the user into the database
"""
import openpyxl
import os
import sys
from datetime import datetime
import json

# Add app path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import (
    User, ExcelProperty, ResidentialComplex, Developer, Manager, 
    District, Building, City, Region, Admin, FavoriteProperty
)

def safe_str(value):
    """Safely convert value to string"""
    if value is None or value == '':
        return None
    return str(value)

def safe_int(value):
    """Safely convert value to integer"""
    if value is None or value == '' or value == 'NULL':
        return None
    try:
        return int(float(value)) if value else None
    except (ValueError, TypeError):
        return None

def safe_float(value):
    """Safely convert value to float"""
    if value is None or value == '' or value == 'NULL':
        return None
    try:
        return float(value) if value else None
    except (ValueError, TypeError):
        return None

def safe_bool(value):
    """Safely convert value to boolean"""
    if value is None or value == '' or value == 'NULL':
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ['true', '1', 'yes', 'да', 'истина']
    return bool(value)

def safe_datetime(value):
    """Safely convert value to datetime"""
    if value is None or value == '' or value == 'NULL':
        return None
    if isinstance(value, datetime):
        return value
    return None

def import_excel_file(file_path, model_class, mapping_func):
    """Generic function to import Excel file into database model"""
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return 0
    
    print(f"Importing {file_path}...")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    imported_count = 0
    errors = []
    
    # Skip header row
    for row_num in range(2, sheet.max_row + 1):
        try:
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col).value
                row_data.append(cell_value)
            
            if not row_data or not any(row_data):  # Skip empty rows
                continue
                
            model_instance = mapping_func(row_data)
            if model_instance:
                db.session.add(model_instance)
                imported_count += 1
                
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            continue
    
    try:
        db.session.commit()
        print(f"Successfully imported {imported_count} records from {file_path}")
    except Exception as e:
        db.session.rollback()
        print(f"Error committing {file_path}: {str(e)}")
        return 0
    
    if errors:
        print(f"Errors encountered: {len(errors)}")
        for error in errors[:5]:  # Show first 5 errors
            print(f"  - {error}")
    
    return imported_count

def map_user_data(row_data):
    """Map Excel row to User model"""
    if len(row_data) < 3:
        return None
    return User(
        email=safe_str(row_data[1]) if len(row_data) > 1 else None,
        phone=safe_str(row_data[2]) if len(row_data) > 2 else None,
        first_name=safe_str(row_data[3]) if len(row_data) > 3 else None,
        last_name=safe_str(row_data[4]) if len(row_data) > 4 else None,
        is_active=safe_bool(row_data[5]) if len(row_data) > 5 else True
    )

def map_excel_property_data(row_data):
    """Map Excel row to ExcelProperty model"""
    if not row_data or not row_data[0]:
        return None
    return ExcelProperty(
        inner_id=safe_int(row_data[0]),
        url=safe_str(row_data[1]) if len(row_data) > 1 else None,
        photos=safe_str(row_data[2]) if len(row_data) > 2 else None,
        address_display_name=safe_str(row_data[15]) if len(row_data) > 15 else None,
        complex_name=safe_str(row_data[18]) if len(row_data) > 18 else None,
        price=safe_int(row_data[58]) if len(row_data) > 58 else None,
        rooms_count=safe_int(row_data[60]) if len(row_data) > 60 else None,
        area=safe_float(row_data[65]) if len(row_data) > 65 else None,
        floor=safe_int(row_data[66]) if len(row_data) > 66 else None,
        floors_count=safe_int(row_data[67]) if len(row_data) > 67 else None
    )

def map_residential_complex_data(row_data):
    """Map Excel row to ResidentialComplex model"""
    if len(row_data) < 2:
        return None
    return ResidentialComplex(
        name=safe_str(row_data[1]) if len(row_data) > 1 else None,
        description=safe_str(row_data[2]) if len(row_data) > 2 else None,
        address=safe_str(row_data[3]) if len(row_data) > 3 else None,
        developer_id=safe_int(row_data[4]) if len(row_data) > 4 else None,
        district_id=safe_int(row_data[5]) if len(row_data) > 5 else None
    )

def map_developer_data(row_data):
    """Map Excel row to Developer model"""
    if len(row_data) < 2:
        return None
    return Developer(
        name=safe_str(row_data[1]) if len(row_data) > 1 else None,
        description=safe_str(row_data[2]) if len(row_data) > 2 else None,
        phone=safe_str(row_data[3]) if len(row_data) > 3 else None,
        email=safe_str(row_data[4]) if len(row_data) > 4 else None,
        website=safe_str(row_data[5]) if len(row_data) > 5 else None
    )

def map_manager_data(row_data):
    """Map Excel row to Manager model"""
    if len(row_data) < 3:
        return None
    return Manager(
        first_name=safe_str(row_data[1]) if len(row_data) > 1 else None,
        last_name=safe_str(row_data[2]) if len(row_data) > 2 else None,
        email=safe_str(row_data[3]) if len(row_data) > 3 else None,
        phone=safe_str(row_data[4]) if len(row_data) > 4 else None,
        is_active=safe_bool(row_data[5]) if len(row_data) > 5 else True
    )

def map_district_data(row_data):
    """Map Excel row to District model"""
    if len(row_data) < 2:
        return None
    return District(
        name=safe_str(row_data[1]) if len(row_data) > 1 else None,
        description=safe_str(row_data[2]) if len(row_data) > 2 else None,
        city_id=safe_int(row_data[3]) if len(row_data) > 3 else None
    )

def comprehensive_import():
    """Import all Excel files"""
    print("=== COMPREHENSIVE EXCEL IMPORT ===")
    
    total_imported = 0
    
    with app.app_context():
        # Define import mappings
        imports = [
            ('attached_assets/users (7)_1756658357102.xlsx', User, map_user_data),
            ('attached_assets/excel_properties_1756658357106.xlsx', ExcelProperty, map_excel_property_data),
            ('attached_assets/residential_complexes (6)_1756658357103.xlsx', ResidentialComplex, map_residential_complex_data),
            ('attached_assets/developers (7)_1756658357106.xlsx', Developer, map_developer_data),
            ('attached_assets/managers (6)_1756658357103.xlsx', Manager, map_manager_data),
            ('attached_assets/districts (5)_1756658357105.xlsx', District, map_district_data),
        ]
        
        for file_path, model_class, mapping_func in imports:
            try:
                count = import_excel_file(file_path, model_class, mapping_func)
                total_imported += count
            except Exception as e:
                print(f"Error importing {file_path}: {str(e)}")
                continue
        
        print(f"\n=== IMPORT COMPLETED ===")
        print(f"Total records imported: {total_imported}")
        
        # Show database statistics
        print("\n=== DATABASE STATISTICS ===")
        stats = [
            ('Users', db.session.query(User).count()),
            ('Properties', db.session.query(ExcelProperty).count()),
            ('Residential Complexes', db.session.query(ResidentialComplex).count()),
            ('Developers', db.session.query(Developer).count()),
            ('Managers', db.session.query(Manager).count()),
            ('Districts', db.session.query(District).count()),
        ]
        
        for name, count in stats:
            print(f"{name}: {count}")

if __name__ == '__main__':
    comprehensive_import()