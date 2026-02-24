"""
Импорт всех 77 столбцов из Excel файла в таблицу excel_properties
Этот скрипт сохраняет каждый столбец Excel без потери данных
"""
import openpyxl
from datetime import datetime
import json
import sys
import os

# Добавим путь к нашему приложению
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import ExcelProperty

def parse_boolean(value):
    """Конвертация строки в boolean"""
    if value is None or value == '' or value == 'NULL':
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ['true', '1', 'yes', 'да']
    return bool(value)

def parse_json_safe(value):
    """Безопасная конвертация в JSON строку"""
    if value is None or value == '' or value == 'NULL':
        return None
    if isinstance(value, str) and (value.startswith('[') or value.startswith('{')):
        return value  # Уже JSON
    if isinstance(value, list):
        return json.dumps(value)
    return str(value)

def parse_datetime(value):
    """Конвертация строки в datetime"""
    if value is None or value == '' or value == 'NULL':
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            # Попробуем разные форматы datetime
            if 'T' in value:
                return datetime.fromisoformat(value.replace('Z', '+00:00'))
            else:
                return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except:
            return None
    return None

def import_excel_data():
    """Импорт всех 455 записей из Excel в БД"""
    print("=== ИМПОРТ ВСЕХ 77 СТОЛБЦОВ EXCEL ДАННЫХ ===")
    
    # Открываем Excel файл
    excel_file = 'attached_assets/Сочи_1756309636907.xlsx'
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.active
    
    print(f"Найден файл с {sheet.max_row} строками и {sheet.max_column} столбцами")
    
    # Очистим существующие данные
    with app.app_context():
        db.session.query(ExcelProperty).delete()
        db.session.commit()
        print("Очистили существующие данные в excel_properties")
    
    imported_count = 0
    errors = []
    
    # Импортируем каждую строку
    for row_num in range(2, sheet.max_row + 1):  # Начинаем с 2-й строки (пропускаем заголовки)
        try:
            # Читаем все значения строки
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col).value
                row_data.append(cell_value)
            
            # Создаем объект ExcelProperty со всеми 77 полями
            excel_property = ExcelProperty(
                # Основные данные (столбцы 1-3)
                inner_id=row_data[0],  # inner_id
                url=row_data[1] if row_data[1] else None,  # url
                photos=parse_json_safe(row_data[2]),  # photos
                
                # Адресная информация (столбцы 4-17)
                address_id=row_data[3] if row_data[3] else None,  # address_id
                address_guid=row_data[4] if row_data[4] else None,  # address_guid
                address_kind=row_data[5] if row_data[5] else None,  # address_kind
                address_name=row_data[6] if row_data[6] else None,  # address_name
                address_subways=parse_json_safe(row_data[7]),  # address_subways
                address_locality_id=row_data[8] if row_data[8] else None,  # address_locality_id
                address_locality_kind=row_data[9] if row_data[9] else None,  # address_locality_kind
                address_locality_name=row_data[10] if row_data[10] else None,  # address_locality_name
                address_locality_subkind=row_data[11] if row_data[11] else None,  # address_locality_subkind
                address_locality_display_name=row_data[12] if row_data[12] else None,  # address_locality_display_name
                address_position_lat=row_data[13] if row_data[13] else None,  # address_position_lat
                address_position_lon=row_data[14] if row_data[14] else None,  # address_position_lon
                address_display_name=row_data[15] if row_data[15] else None,  # address_display_name
                address_short_display_name=row_data[16] if row_data[16] else None,  # address_short_display_name
                
                # Данные о ЖК (столбцы 18-32)
                complex_id=row_data[17] if row_data[17] else None,  # complex_id
                complex_name=row_data[18] if row_data[18] else None,  # complex_name
                complex_phone=row_data[19] if row_data[19] else None,  # complex_phone
                complex_building_id=row_data[20] if row_data[20] else None,  # complex_building_id
                complex_building_name=row_data[21] if row_data[21] else None,  # complex_building_name
                complex_building_released=parse_boolean(row_data[22]),  # complex_building_released
                complex_building_is_unsafe=parse_boolean(row_data[23]),  # complex_building_is_unsafe
                complex_building_accreditation=parse_boolean(row_data[24]),  # complex_building_accreditation
                complex_building_end_build_year=row_data[25] if row_data[25] else None,  # complex_building_end_build_year
                complex_building_complex_product=parse_boolean(row_data[26]),  # complex_building_complex_product
                complex_building_end_build_quarter=row_data[27] if row_data[27] else None,  # complex_building_end_build_quarter
                complex_building_has_green_mortgage=parse_boolean(row_data[28]),  # complex_building_has_green_mortgage
                complex_min_rate=row_data[29] if row_data[29] else None,  # complex_min_rate
                complex_sales_phone=parse_json_safe(row_data[30]),  # complex_sales_phone
                complex_sales_address=row_data[31] if row_data[31] else None,  # complex_sales_address
                
                # Характеристики ЖК (столбцы 33-53)
                complex_object_class_id=row_data[32] if row_data[32] else None,  # complex_object_class_id
                complex_object_class_display_name=row_data[33] if row_data[33] else None,  # complex_object_class_display_name
                complex_has_big_check=parse_boolean(row_data[34]),  # complex_has_big_check
                complex_end_build_year=row_data[35] if row_data[35] else None,  # complex_end_build_year
                complex_financing_sber=parse_boolean(row_data[36]),  # complex_financing_sber
                complex_telephony_b_number=row_data[37] if row_data[37] else None,  # complex_telephony_b_number
                complex_telephony_r_number=row_data[38] if row_data[38] else None,  # complex_telephony_r_number
                complex_with_renovation=parse_boolean(row_data[39]),  # complex_with_renovation
                complex_first_build_year=row_data[40] if row_data[40] else None,  # complex_first_build_year
                complex_start_build_year=row_data[41] if row_data[41] else None,  # complex_start_build_year
                complex_end_build_quarter=row_data[42] if row_data[42] else None,  # complex_end_build_quarter
                complex_has_accreditation=parse_boolean(row_data[43]),  # complex_has_accreditation
                complex_has_approve_flats=parse_boolean(row_data[44]),  # complex_has_approve_flats
                complex_mortgage_tranches=parse_boolean(row_data[45]),  # complex_mortgage_tranches
                complex_has_green_mortgage=parse_boolean(row_data[46]),  # complex_has_green_mortgage
                complex_phone_substitution=row_data[47] if row_data[47] else None,  # complex_phone_substitution
                complex_show_contact_block=parse_boolean(row_data[48]),  # complex_show_contact_block
                complex_first_build_quarter=row_data[49] if row_data[49] else None,  # complex_first_build_quarter
                complex_start_build_quarter=row_data[50] if row_data[50] else None,  # complex_start_build_quarter
                complex_has_mortgage_subsidy=parse_boolean(row_data[51]),  # complex_has_mortgage_subsidy
                complex_has_government_program=parse_boolean(row_data[52]),  # complex_has_government_program
                
                # Условия и параметры (столбцы 54-61)
                min_rate=row_data[53] if row_data[53] else None,  # min_rate
                trade_in=parse_boolean(row_data[54]),  # trade_in
                deal_type=row_data[55] if row_data[55] else None,  # deal_type
                developer_id=row_data[56] if row_data[56] else None,  # developer_id
                developer_name=row_data[57] if row_data[57] else None,  # developer_name
                developer_site=row_data[58] if row_data[58] else None,  # developer_site
                developer_holding_id=row_data[59] if row_data[59] else None,  # developer_holding_id
                is_auction=parse_boolean(row_data[60]),  # is_auction
                
                # Цены (столбцы 62-66)
                price=row_data[61] if row_data[61] else None,  # price
                max_price=row_data[62] if row_data[62] else None,  # max_price
                min_price=row_data[63] if row_data[63] else None,  # min_price
                square_price=row_data[64] if row_data[64] else None,  # square_price
                mortgage_price=row_data[65] if row_data[65] else None,  # mortgage_price
                
                # Характеристики квартиры (столбцы 67-74)
                renovation_type=row_data[66] if row_data[66] else None,  # renovation_type
                renovation_display_name=row_data[67] if row_data[67] else None,  # renovation_display_name
                description=row_data[68] if row_data[68] else None,  # description
                object_area=row_data[69] if row_data[69] else None,  # object_area
                object_rooms=row_data[70] if row_data[70] else None,  # object_rooms
                object_max_floor=row_data[71] if row_data[71] else None,  # object_max_floor
                object_min_floor=row_data[72] if row_data[72] else None,  # object_min_floor
                object_is_apartment=parse_boolean(row_data[73]),  # object_is_apartment
                
                # Метаданные (столбцы 75-77)
                published_dt=parse_datetime(row_data[74]),  # published_dt
                chat_available=parse_boolean(row_data[75]),  # chat_available
                placement_type=parse_json_safe(row_data[76]) if len(row_data) > 76 else None,  # placement_type
            )
            
            with app.app_context():
                db.session.add(excel_property)
                db.session.commit()
            
            imported_count += 1
            if imported_count % 50 == 0:
                print(f"Импортировано {imported_count} записей...")
                
        except Exception as e:
            error_msg = f"Ошибка в строке {row_num}: {e}"
            errors.append(error_msg)
            print(error_msg)
            continue
    
    print(f"\n=== РЕЗУЛЬТАТЫ ИМПОРТА ===")
    print(f"Успешно импортировано: {imported_count} записей")
    print(f"Ошибок: {len(errors)}")
    
    if errors:
        print("\nОшибки:")
        for error in errors[:10]:  # Показываем первые 10 ошибок
            print(f"  - {error}")
        if len(errors) > 10:
            print(f"  ... и еще {len(errors) - 10} ошибок")
    
    # Проверяем результат
    with app.app_context():
        total_count = db.session.query(ExcelProperty).count()
        sample = db.session.query(ExcelProperty).first()
        
        print(f"\n=== ПРОВЕРКА ===")
        print(f"Всего записей в БД: {total_count}")
        if sample:
            print(f"Пример записи: inner_id={sample.inner_id}, complex_name={sample.complex_name}, price={sample.price}")

if __name__ == "__main__":
    import_excel_data()