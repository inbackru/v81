#!/usr/bin/env python3
"""
Полный импорт всех 454 объектов из Excel со всеми 77 колонками
"""
import openpyxl
import os
from app import app, db
from sqlalchemy import text
import json

def import_all_properties():
    """Импорт всех объектов из Excel"""
    
    excel_file = 'attached_assets/Сочи_1759349636274.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Файл {excel_file} не найден!")
        return False
    
    print(f"📂 Загрузка файла: {excel_file}")
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Получаем заголовки
        headers = [cell.value for cell in ws[1]]
        print(f"✅ Колонок в Excel: {len(headers)}")
        print(f"✅ Строк данных: {ws.max_row - 1}")
        
        with app.app_context():
            # Сначала очищаем таблицу
            print("\n🗑️  Очистка таблицы excel_properties...")
            db.session.execute(text("DELETE FROM excel_properties"))
            db.session.commit()
            
            # Проверяем, что таблица пустая
            count = db.session.execute(text("SELECT COUNT(*) FROM excel_properties")).scalar()
            print(f"✅ Таблица очищена (осталось записей: {count})")
            
            inserted_count = 0
            error_count = 0
            skipped_duplicates = 0
            seen_ids = set()  # Для отслеживания дубликатов
            
            print("\n📥 Импорт данных...")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Преобразуем строку в словарь
                    data = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            value = row[idx]
                            # Преобразуем булевы значения
                            if isinstance(value, str):
                                if value.lower() == 'true':
                                    value = True
                                elif value.lower() == 'false':
                                    value = False
                            data[header] = value
                    
                    inner_id = data.get('inner_id')
                    if not inner_id:
                        continue
                    
                    # Пропускаем дубликаты (берем только первое вхождение)
                    if inner_id in seen_ids:
                        skipped_duplicates += 1
                        continue
                    seen_ids.add(inner_id)
                    
                    # SQL для вставки с ALL колонками
                    insert_sql = text("""
                        INSERT INTO excel_properties (
                            inner_id, url, photos,
                            address_id, address_guid, address_kind, address_name, address_subways,
                            address_locality_id, address_locality_kind, address_locality_name,
                            address_locality_subkind, address_locality_display_name,
                            address_position_lat, address_position_lon,
                            address_display_name, address_short_display_name,
                            complex_id, complex_name, complex_phone,
                            complex_building_id, complex_building_name, complex_building_released,
                            complex_building_is_unsafe, complex_building_accreditation,
                            complex_building_end_build_year, complex_building_complex_product,
                            complex_building_end_build_quarter, complex_building_has_green_mortgage,
                            complex_min_rate, complex_sales_phone, complex_sales_address,
                            complex_object_class_id, complex_object_class_display_name,
                            complex_has_big_check, complex_end_build_year, complex_financing_sber,
                            complex_telephony_b_number, complex_telephony_r_number,
                            complex_with_renovation, complex_first_build_year, complex_start_build_year,
                            complex_end_build_quarter, complex_has_accreditation,
                            complex_has_approve_flats, complex_mortgage_tranches,
                            complex_has_green_mortgage, complex_phone_substitution,
                            complex_show_contact_block, complex_first_build_quarter,
                            complex_start_build_quarter, complex_has_mortgage_subsidy,
                            complex_has_government_program,
                            min_rate, trade_in, deal_type,
                            developer_id, developer_name, developer_site, developer_holding_id,
                            is_auction, price, max_price, min_price, square_price, mortgage_price,
                            renovation_type, renovation_display_name, description,
                            object_area, object_rooms, object_max_floor, object_min_floor,
                            object_is_apartment, published_dt, chat_available, placement_type
                        ) VALUES (
                            :inner_id, :url, :photos,
                            :address_id, :address_guid, :address_kind, :address_name, :address_subways,
                            :address_locality_id, :address_locality_kind, :address_locality_name,
                            :address_locality_subkind, :address_locality_display_name,
                            :address_position_lat, :address_position_lon,
                            :address_display_name, :address_short_display_name,
                            :complex_id, :complex_name, :complex_phone,
                            :complex_building_id, :complex_building_name, :complex_building_released,
                            :complex_building_is_unsafe, :complex_building_accreditation,
                            :complex_building_end_build_year, :complex_building_complex_product,
                            :complex_building_end_build_quarter, :complex_building_has_green_mortgage,
                            :complex_min_rate, :complex_sales_phone, :complex_sales_address,
                            :complex_object_class_id, :complex_object_class_display_name,
                            :complex_has_big_check, :complex_end_build_year, :complex_financing_sber,
                            :complex_telephony_b_number, :complex_telephony_r_number,
                            :complex_with_renovation, :complex_first_build_year, :complex_start_build_year,
                            :complex_end_build_quarter, :complex_has_accreditation,
                            :complex_has_approve_flats, :complex_mortgage_tranches,
                            :complex_has_green_mortgage, :complex_phone_substitution,
                            :complex_show_contact_block, :complex_first_build_quarter,
                            :complex_start_build_quarter, :complex_has_mortgage_subsidy,
                            :complex_has_government_program,
                            :min_rate, :trade_in, :deal_type,
                            :developer_id, :developer_name, :developer_site, :developer_holding_id,
                            :is_auction, :price, :max_price, :min_price, :square_price, :mortgage_price,
                            :renovation_type, :renovation_display_name, :description,
                            :object_area, :object_rooms, :object_max_floor, :object_min_floor,
                            :object_is_apartment, :published_dt, :chat_available, :placement_type
                        )
                    """)
                    
                    db.session.execute(insert_sql, data)
                    inserted_count += 1
                    
                    if inserted_count % 100 == 0:
                        print(f"   Импортировано: {inserted_count}")
                        db.session.commit()
                        
                except Exception as e:
                    error_count += 1
                    if error_count <= 5:  # Показываем первые 5 ошибок
                        print(f"❌ Ошибка в строке {row_idx} (ID: {data.get('inner_id', 'unknown')}): {e}")
                    continue
            
            # Финальный коммит
            db.session.commit()
            
            print()
            print("=" * 70)
            print("✅ Импорт завершен!")
            print(f"   Добавлено записей: {inserted_count}")
            print(f"   Пропущено дубликатов: {skipped_duplicates}")
            print(f"   Ошибок: {error_count}")
            print("=" * 70)
            
            # Статистика
            stats = db.session.execute(text("""
                SELECT 
                    COUNT(*) as total,
                    COUNT(CASE WHEN address_position_lat IS NOT NULL THEN 1 END) as with_coords,
                    COUNT(complex_building_end_build_year) as with_deadline,
                    COUNT(DISTINCT complex_name) as unique_complexes,
                    COUNT(DISTINCT developer_name) as unique_developers
                FROM excel_properties
            """)).fetchone()
            
            print()
            print("📊 Статистика базы данных:")
            print(f"   Всего объектов: {stats[0]}")
            print(f"   С координатами (видны на карте): {stats[1]}")
            print(f"   Со сроками сдачи: {stats[2]}")
            print(f"   Уникальных ЖК: {stats[3]}")
            print(f"   Уникальных застройщиков: {stats[4]}")
            
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Полный импорт всех объектов из Excel")
    print("=" * 70)
    print()
    
    success = import_all_properties()
    
    if success:
        print()
        print("✅ Импорт успешно завершен!")
        print("🔄 Перезапустите приложение для применения изменений")
    else:
        print()
        print("❌ Импорт завершился с ошибками")
