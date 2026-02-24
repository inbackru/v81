"""
Simple data import script matching actual database schema
Imports Excel data into the existing database structure
"""
import openpyxl
import sys
import os
from app import app, db
from sqlalchemy import text

def import_from_excel(file_path, table_name, column_mapping):
    """Import Excel data into specified table"""
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return 0
    
    print(f"Importing {file_path} into {table_name}...")
    
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
    
    imported_count = 0
    
    for row_num in range(2, sheet.max_row + 1):
        try:
            row_data = {}
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col).value
                if headers[col-1] in column_mapping:
                    db_column = column_mapping[headers[col-1]]
                    row_data[db_column] = cell_value
            
            if not row_data:
                continue
            
            # Create INSERT statement
            columns = list(row_data.keys())
            values = list(row_data.values())
            
            # Clean None values
            clean_data = {}
            for k, v in row_data.items():
                if v is not None and v != '':
                    clean_data[k] = v
            
            if clean_data:
                columns_str = ', '.join(clean_data.keys())
                placeholders = ', '.join([f":{k}" for k in clean_data.keys()])
                
                sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"
                
                try:
                    db.session.execute(text(sql), clean_data)
                    imported_count += 1
                except Exception as e:
                    if "duplicate key value" not in str(e).lower():
                        print(f"Error inserting row {row_num}: {e}")
                    continue
        except Exception as e:
            print(f"Error processing row {row_num}: {e}")
            continue
    
    try:
        db.session.commit()
        print(f"Successfully imported {imported_count} records into {table_name}")
    except Exception as e:
        db.session.rollback()
        print(f"Error committing to {table_name}: {e}")
        return 0
    
    return imported_count

def main():
    """Main import function"""
    print("=== SIMPLE DATA IMPORT ===")
    
    with app.app_context():
        total_imported = 0
        
        # Import users
        user_mapping = {
            'email': 'email',
            'phone': 'phone',
            'full_name': 'full_name',
            'name': 'full_name',
            'first_name': 'full_name'
        }
        
        users_file = 'attached_assets/users (7)_1756658357102.xlsx'
        total_imported += import_from_excel(users_file, 'users', user_mapping)
        
        # Import districts
        district_mapping = {
            'name': 'name',
            'slug': 'slug'
        }
        
        districts_file = 'attached_assets/districts (5)_1756658357105.xlsx'
        total_imported += import_from_excel(districts_file, 'districts', district_mapping)
        
        # Import developers
        developer_mapping = {
            'name': 'name',
            'description': 'description',
            'phone': 'phone',
            'email': 'email',
            'website': 'website'
        }
        
        developers_file = 'attached_assets/developers (7)_1756658357106.xlsx'
        total_imported += import_from_excel(developers_file, 'developers', developer_mapping)
        
        # Import residential complexes
        complex_mapping = {
            'name': 'name',
            'description': 'description',
            'address': 'address'
        }
        
        complexes_file = 'attached_assets/residential_complexes (6)_1756658357103.xlsx'
        total_imported += import_from_excel(complexes_file, 'residential_complexes', complex_mapping)
        
        # Import managers
        manager_mapping = {
            'first_name': 'first_name',
            'last_name': 'last_name',
            'email': 'email',
            'phone': 'phone'
        }
        
        managers_file = 'attached_assets/managers (6)_1756658357103.xlsx'
        total_imported += import_from_excel(managers_file, 'managers', manager_mapping)
        
        print(f"\n✅ Import completed! Total records imported: {total_imported}")
        
        # Show counts
        print("\n=== DATABASE COUNTS ===")
        tables = ['users', 'districts', 'developers', 'residential_complexes', 'managers']
        for table in tables:
            result = db.session.execute(text(f"SELECT COUNT(*) FROM {table}"))
            count = result.scalar()
            print(f"{table}: {count}")

if __name__ == '__main__':
    main()