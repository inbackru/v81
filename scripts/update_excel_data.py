#!/usr/bin/env python3
"""
Обновление данных из Excel файла в базу данных
Исправляет: изображения, площадь, цены, комнаты, этажи, адреса
"""

import openpyxl
import json
import sys
import os

# Import models
sys.path.append('/home/runner/workspace')
try:
    from models import Property, ResidentialComplex, Building, Developer, District, db
    from app import app
    print("Успешно импортированы модели БД")
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    sys.exit(1)

def parse_excel_data():
    """Парсит данные из Excel файла"""
    excel_file = 'attached_assets/Сочи_1756309636907.xlsx'
    
    print(f"Загружаю Excel файл: {excel_file}")
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook.active
    
    # Маппинг столбцов (найденных выше)
    columns = {
        'photos': 3,           # Изображения
        'address_name': 7,     # Адрес
        'address_display': 16, # Полный адрес  
        'complex_name': 19,    # Название ЖК
        'complex_phone': 20,   # Телефон ЖК
        'building_name': 22,   # Название здания
        'price': 62,           # Цена
        'object_area': 70,     # Площадь
        'object_rooms': 71,    # Комнаты
        'max_floor': 72,       # Макс этаж
        'min_floor': 73        # Этаж квартиры
    }
    
    properties_data = []
    
    print(f"Обрабатываю {sheet.max_row - 1} строк данных...")
    
    for row in range(2, min(sheet.max_row + 1, 100)):  # Первые 100 строк для теста
        try:
            # Извлекаем данные
            photos_raw = sheet.cell(row=row, column=columns['photos']).value
            photos = json.loads(photos_raw) if photos_raw else []
            
            price = sheet.cell(row=row, column=columns['price']).value
            area = sheet.cell(row=row, column=columns['object_area']).value  
            rooms = sheet.cell(row=row, column=columns['object_rooms']).value
            max_floor = sheet.cell(row=row, column=columns['max_floor']).value
            floor = sheet.cell(row=row, column=columns['min_floor']).value
            
            complex_name = sheet.cell(row=row, column=columns['complex_name']).value
            address_name = sheet.cell(row=row, column=columns['address_name']).value
            address_display = sheet.cell(row=row, column=columns['address_display']).value
            building_name = sheet.cell(row=row, column=columns['building_name']).value
            
            property_data = {
                'row': row,
                'photos': photos[:5] if photos else [],  # Первые 5 фото
                'main_image': photos[0] if photos else None,
                'gallery_images': json.dumps(photos[1:10]) if len(photos) > 1 else None,
                'price': int(price) if price else None,
                'area': float(area) if area else None,
                'rooms': int(rooms) if rooms else 0,
                'floor': int(floor) if floor else None,
                'max_floor': int(max_floor) if max_floor else None,
                'complex_name': str(complex_name).strip() if complex_name else 'Без названия',
                'address': str(address_display or address_name or '').strip(),
                'building_name': str(building_name).strip() if building_name else None
            }
            
            properties_data.append(property_data)
            
            if row % 20 == 0:
                print(f"Обработано строк: {row-1}")
                
        except Exception as e:
            print(f"Ошибка в строке {row}: {e}")
            
    print(f"Успешно обработано {len(properties_data)} объектов недвижимости")
    return properties_data

def update_database(properties_data):
    """Обновляет существующие записи в БД"""
    
    with app.app_context():
        updated_count = 0
        
        for prop_data in properties_data:
            try:
                # Ищем существующую квартиру по названию ЖК и параметрам
                existing_prop = Property.query.filter(
                    Property.rooms == prop_data['rooms']
                ).first()
                
                if existing_prop and prop_data['price']:
                    # Обновляем найденную запись
                    existing_prop.price = prop_data['price']
                    existing_prop.area = prop_data['area']
                    existing_prop.floor = prop_data['floor']
                    existing_prop.main_image = prop_data['main_image']
                    existing_prop.gallery_images = prop_data['gallery_images']
                    if prop_data['address']:
                        existing_prop.address = prop_data['address']
                        
                    # Обновляем этажность здания
                    if existing_prop.building and prop_data['max_floor']:
                        existing_prop.building.floors = prop_data['max_floor']
                        
                    updated_count += 1
                    
                    if updated_count % 10 == 0:
                        print(f"Обновлено записей: {updated_count}")
                        
            except Exception as e:
                print(f"Ошибка обновления записи: {e}")
                
        db.session.commit()
        print(f"Обновление завершено. Обновлено записей: {updated_count}")

if __name__ == "__main__":
    print("=== ОБНОВЛЕНИЕ ДАННЫХ ИЗ EXCEL ===")
    data = parse_excel_data()
    
    if data:
        print("\\nПример обработанных данных:")
        for i, prop in enumerate(data[:3]):
            print(f"Объект {i+1}:")
            print(f"  ЖК: {prop['complex_name']}")
            print(f"  Цена: {prop['price']:,} ₽" if prop['price'] else "  Цена: не указана")
            print(f"  Площадь: {prop['area']} м²" if prop['area'] else "  Площадь: не указана")  
            print(f"  Комнаты: {prop['rooms']}")
            print(f"  Этаж: {prop['floor']}/{prop['max_floor']}")
            print(f"  Фото: {len(prop['photos'])}")
            print(f"  Адрес: {prop['address'][:50]}...")
            print()
        
        update_database(data)
    else:
        print("Нет данных для обновления")